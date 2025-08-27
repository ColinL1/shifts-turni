from openpyxl import load_workbook
import csv

# Read the Excel file and convert to CSV
wb = load_workbook('../ostardo_turni.xlsx')
ws = wb.active

# Write to CSV
with open('../ostardo_turni_improved.csv', 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)

print("Created CSV version: ../ostardo_turni_improved.csv")

# Print the contents
print("\nPreview of the data:")
with open('../ostardo_turni_improved.csv', 'r', encoding='utf-8') as f:
    print(f.read())
