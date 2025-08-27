from openpyxl import load_workbook
import csv

# Read the Excel file and show all sheets
wb = load_workbook('../ostardo_turni.xlsx')

print("Excel file contains the following sheets:")
for sheet_name in wb.sheetnames:
    print(f"- {sheet_name}")

print("\n" + "="*50)

# Display each sheet
for sheet_name in wb.sheetnames:
    print(f"\nSheet: {sheet_name}")
    print("-" * (len(sheet_name) + 7))
    
    ws = wb[sheet_name]
    
    # Convert sheet to CSV format for display
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):  # Skip empty rows
            row_data = [str(cell) if cell is not None else '' for cell in row]
            print(','.join(row_data))
    
    print()  # Empty line between sheets
