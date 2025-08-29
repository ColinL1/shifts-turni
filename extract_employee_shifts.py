import os
import re
import sys
import time
from datetime import datetime, timedelta
from docx import Document
from openpyxl import Workbook
from typing import Dict, Any, List, Tuple

# Folder containing the .docx files
TURNI_FOLDER = 'turni'
OUTPUT_XLSX = 'employee_shifts.xlsx'

# Helper to extract date range from filename (e.g., '57. 25/11/24 - 29/11/24.docx' or '59. 09:12:24 - 13:12:24.docx' or '55. 11:11 - 15:11.docx')
def extract_date_range_from_filename(filename):
    # First try the new format with forward slashes: "25/11/24 - 29/11/24"
    match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{2,4})', filename)
    if match:
        start_day, start_month, start_year, end_day, end_month, end_year = match.groups()
        # Convert 2-digit years to 4-digit (24 -> 2024)
        start_year = int(start_year)
        end_year = int(end_year)
        if start_year < 100:
            start_year += 2000
        if end_year < 100:
            end_year += 2000
        return int(start_day), int(start_month), int(end_day), int(end_month), start_year, end_year
    
    # Try format with colons and year: "09:12:24 - 13:12:24" (day:month:year)
    match = re.search(r'(\d{1,2}):(\d{1,2}):(\d{2,4})\s*-\s*(\d{1,2}):(\d{1,2}):(\d{2,4})', filename)
    if match:
        start_day, start_month, start_year, end_day, end_month, end_year = match.groups()
        # Convert 2-digit years to 4-digit (24 -> 2024, 25 -> 2025)
        start_year = int(start_year)
        end_year = int(end_year)
        if start_year < 100:
            start_year += 2000
        if end_year < 100:
            end_year += 2000
        return int(start_day), int(start_month), int(end_day), int(end_month), start_year, end_year
    
    # Fall back to old format: "11:11 - 15:11" (day:month without year)
    match = re.search(r'(\d{1,2}):(\d{1,2})\s*-\s*(\d{1,2}):(\d{1,2})', filename)
    if match:
        start_day, start_month, end_day, end_month = match.groups()
        return int(start_day), int(start_month), int(end_day), int(end_month), None, None
    
    return None, None, None, None, None, None

# Helper to get the year based on the month (November 2024 to 2025)
def get_year_for_month(month):
    # Assuming the schedule starts in November 2024 and continues into 2025
    # November and December are 2024, January onwards are 2025
    if month >= 11:  # November, December
        return 2024
    else:  # January onwards
        return 2025

# Helper to get week dates from the date range
def get_week_dates_from_range(start_day, start_month, end_day, end_month, start_year=None, end_year=None):
    # Use provided years if available, otherwise use the old logic
    if start_year is None:
        start_year = get_year_for_month(start_month)
    if end_year is None:
        end_year = get_year_for_month(end_month)
    
    start_date = datetime(start_year, start_month, start_day)
    end_date = datetime(end_year, end_month, end_day)
    
    # Generate dates for the work week (Monday to Friday)
    week_dates = []
    current_date = start_date
    
    while current_date <= end_date:
        # Only include weekdays (Monday=0 to Friday=4)
        if current_date.weekday() < 5:
            week_dates.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)
    
    return week_dates

def add_days_to_date(date_str, days):
    """Add days to a date string and return the new date string"""
    if not date_str:
        return ''
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        new_date = date_obj + timedelta(days=days)
        return new_date.strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        return ''

############################
# Performance Helpers
############################

def get_docx_files(folder: str) -> List[str]:
    """Fast directory scan using scandir."""
    files: List[str] = []
    try:
        with os.scandir(folder) as it:
            for entry in it:
                if not entry.is_file():
                    continue
                name = entry.name
                if name.endswith('.docx') and not name.startswith('~$') and not re.match(r'^\d+_\.', name):
                    files.append(entry.path)
    except FileNotFoundError:
        return []
    return files

# Simple in-process cache: { filepath: { 'mtime': float, 'parsed': <Document>, 'tables': preprocessed } }
DOCX_CACHE: Dict[str, Dict[str, Any]] = {}

def _preprocess_tables(doc: Document) -> List[Tuple[List[str], List[List[str]]]]:
    """Return a lightweight representation of each table: (header_cells, row_cells_matrix)."""
    rep = []
    for table in doc.tables:
        try:
            if not table.rows:
                continue
            header = [c.text.strip() for c in table.rows[0].cells]
            body = []
            for row in table.rows[1:]:
                body.append([c.text.strip() for c in row.cells])
            rep.append((header, body))
        except Exception:
            continue
    return rep

def load_doc_cached(filepath: str) -> Dict[str, Any]:
    """Load a docx file with caching based on mtime."""
    try:
        mtime = os.path.getmtime(filepath)
    except FileNotFoundError:
        return {}
    cached = DOCX_CACHE.get(filepath)
    if cached and cached.get('mtime') == mtime:
        return cached
    # (Re)load
    doc = Document(filepath)
    data = {
        'mtime': mtime,
        'tables': _preprocess_tables(doc)
    }
    DOCX_CACHE[filepath] = data
    return data

# Main extraction logic
def extract_employee_shifts(employee_name: str):
    """Extract shifts for an employee with caching and minimal re-parsing."""
    t0 = time.time()
    target_lc = employee_name.lower()
    results: List[Dict[str, str]] = []
    files = get_docx_files(TURNI_FOLDER)
    for filepath in files:
        filename = os.path.basename(filepath)
        # Extract date range from filename
        start_day, start_month, end_day, end_month, start_year, end_year = extract_date_range_from_filename(filename)
        if not all([start_day, start_month, end_day, end_month]):
            continue
        week_dates = get_week_dates_from_range(start_day, start_month, end_day, end_month, start_year, end_year)
        data = load_doc_cached(filepath)
        tables = data.get('tables', [])
        for header, body in tables:
            if not header:
                continue
            days = header[1:]
            for row in body:
                if not row:
                    continue
                shift_type = row[0].strip()
                # Skip rows with no employee cells
                for i, cell_text in enumerate(row[1:]):
                    cell_norm = cell_text.lower()
                    if target_lc in cell_norm:
                        if 'assenti' in shift_type.lower():
                            continue
                        day_name = days[i] if i < len(days) else f'Day{i+1}'
                        date_str = week_dates[i] if i < len(week_dates) else ''
                        result_entry = {
                            'File': filename,
                            'Data': date_str,
                            'Giorno': day_name,
                            'Turno': shift_type
                        }
                        results.append(result_entry)
                        if 'guardia' in shift_type.lower() and day_name.lower() == 'venerdì' and date_str:
                            base_dt = datetime.strptime(date_str, '%Y-%m-%d')
                            for add_days, extra_day in [(1, 'Sabato'), (2, 'Domenica')]:
                                extra_dt = base_dt + timedelta(days=add_days)
                                results.append({
                                    'File': filename,
                                    'Data': extra_dt.strftime('%Y-%m-%d'),
                                    'Giorno': extra_day,
                                    'Turno': shift_type
                                })
    elapsed = (time.time() - t0) * 1000
    # Lightweight performance log to stdout (captured by docker logs)
    print(f"[perf] extract_employee_shifts employee='{employee_name}' files={len(files)} results={len(results)} time_ms={elapsed:.1f}")
    return results

def write_to_xlsx(data, output_path):
    from collections import Counter
    
    # Sort data by date (oldest to newest)
    sorted_data = sorted(data, key=lambda x: x['Data'])
    
    wb = Workbook()
    
    # Sheet 1: All shifts (sorted by date)
    ws1 = wb.active
    ws1.title = 'Tutti i Turni'
    headers = ['File', 'Data', 'Giorno', 'Turno']
    ws1.append(headers)
    for row in sorted_data:
        ws1.append([row[h] for h in headers])
    
    # Sheet 2: Summary count by shift type
    ws2 = wb.create_sheet(title='Riepilogo per Turno')
    shift_counts = Counter(row['Turno'] for row in sorted_data)
    ws2.append(['Tipo di Turno', 'Numero di Volte'])
    for shift_type, count in sorted(shift_counts.items()):
        ws2.append([shift_type, count])
    
    # Sheet 3: Dates grouped by shift type (horizontal layout)
    ws3 = wb.create_sheet(title='Date per Turno')
    
    # Group data by shift type
    shifts_by_type = {}
    for row in sorted_data:
        shift_type = row['Turno']
        if shift_type not in shifts_by_type:
            shifts_by_type[shift_type] = []
        shifts_by_type[shift_type].append(row['Data'])
    
    # Sort shift types and get unique sorted dates for each
    sorted_shift_types = sorted(shifts_by_type.keys())
    for shift_type in sorted_shift_types:
        shifts_by_type[shift_type] = sorted(list(set(shifts_by_type[shift_type])))
    
    # Write headers (shift types)
    for col, shift_type in enumerate(sorted_shift_types, 1):
        ws3.cell(row=1, column=col, value=shift_type)
    
    # Write dates under each shift type column
    for col, shift_type in enumerate(sorted_shift_types, 1):
        dates = shifts_by_type[shift_type]
        for row, date in enumerate(dates, 2):  # Start from row 2
            ws3.cell(row=row, column=col, value=date)
    
    wb.save(output_path)

if __name__ == '__main__':
    # Check if employee name is provided as command line argument
    if len(sys.argv) != 2:
        print("Usage: python extract_employee_shifts.py <employee_name>")
        print("Example: python extract_employee_shifts.py 'John Doe'")
        sys.exit(1)
    
    employee_name = sys.argv[1]
    print(f"Extracting shifts for: {employee_name}")
    
    shifts = extract_employee_shifts(employee_name)
    
    if not shifts:
        print(f"No shifts found for employee: {employee_name}")
        print("Please check the employee name spelling and try again.")
        sys.exit(1)
    
    write_to_xlsx(shifts, OUTPUT_XLSX)
    print(f'Saved {len(shifts)} shifts for {employee_name} to {OUTPUT_XLSX}')
